# # app.py
# import streamlit as st
# import pandas as pd
# import io
# import re
# from datetime import datetime
# import calendar

# # For Excel cell formatting
# from openpyxl.styles import PatternFill

# st.set_page_config(page_title="AWC Multi-file Merger", layout="wide")

import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
import calendar

# For Excel cell formatting
from openpyxl.styles import PatternFill

st.set_page_config(page_title="ICDS AWC Status Tracker", layout="wide")

st.title("🏥 ICDS AWC Status Tracker")
st.markdown("### Multi-file Merger & Compliance Monitor")

st.markdown("""
Upload up to 100 CSV/XLSX files whose filenames contain a date in the format `dd_mm_yyyy`
(or `dd_mm_yyyy(1)` etc). Also upload a single CSV/XLSX file that lists AWC names/codes (one column).
The app will filter rows across all uploaded daily files for AWCs present in that list, add `Date`
and `Day` columns (parsed from filename), and produce a combined file grouped by AWC.
""")

# Color Legend
st.markdown("---")
st.markdown("### 📊 Color Coding Legend")

col_legend1, col_legend2, col_legend3, col_legend4 = st.columns(4)

with col_legend1:
    st.markdown("""
    <div style="background-color: #FFFF00; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #E6E600;">
        <strong style="color: #000;">🟡 YELLOW</strong><br>
        <span style="font-size: 13px; color: #333;">AWC didn't open = 1<br>(Centre Closed)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend2:
    st.markdown("""
    <div style="background-color: #FFA500; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #E69500;">
        <strong style="color: #000;">🟠 ORANGE</strong><br>
        <span style="font-size: 13px; color: #333;">AWC open = 0 & Total HCM = 0<br>(No Hot Cooked Meal)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend3:
    st.markdown("""
    <div style="background-color: #FFC0CB; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FFB0BB;">
        <strong style="color: #000;">🌸 PINK</strong><br>
        <span style="font-size: 13px; color: #333;">Snack ≠ 0 on Mon/Wed/Fri<br>(Snack on Non-Snack Day)</span>
    </div>
    """, unsafe_allow_html=True)

with col_legend4:
    st.markdown("""
    <div style="background-color: #FFC7CE; padding: 15px; border-radius: 8px; text-align: center; border: 2px solid #FFB7BE;">
        <strong style="color: #000;">🔴 RED</strong><br>
        <span style="font-size: 13px; color: #333;">Snack = 0 on Tue/Thu/Sat<br>(Missing Snack on Snack Day)</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")

st.markdown("""
**Processing Rules:**
- **Rule 0:** All Sunday rows are automatically removed from the final output
- **Rule 1:** Yellow highlighting indicates the AWC was closed
- **Rule 2:** Orange highlighting indicates no hot cooked meals were provided despite AWC being open
- **Rule 3:** Pink highlighting indicates snacks were given on days when they shouldn't be (Mon/Wed/Fri)
- **Rule 4:** Red highlighting indicates missing snacks on designated snack days (Tue/Thu/Sat)

*Note: Original data values are preserved; only cell highlighting is applied in the XLSX output.*
""")

# --- Inputs -----------------------------------------------------------------------------
col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("1) Upload daily files (up to 100)")
    uploaded_files = st.file_uploader(
        "Choose CSV or XLSX files (multiple). Filenames must contain date like _dd_mm_yyyy",
        type=["csv", "xlsx"],
        accept_multiple_files=True,
        help="Examples: file_name_ABC_01_01_2025.csv or file_name_DEF_01_01_2025(1).xlsx"
    )

with col2:
    st.subheader("2) Upload AWC list file (single)")
    awc_file = st.file_uploader(
        "AWC list (csv or xlsx) — single file with one column containing AWC names/codes",
        type=["csv", "xlsx"],
        accept_multiple_files=False
    )

st.markdown("---")

# Options
st.sidebar.header("Options")
awc_column_hint = st.sidebar.text_input("If you know the AWC column header, enter it (optional)", "")
output_format = st.sidebar.selectbox("Output format", ["xlsx", "csv"])
sort_by_date = st.sidebar.checkbox("Sort each AWC's rows by Date ascending", value=True)
display_sample = st.sidebar.number_input("Show up to N matching rows in preview", min_value=0, max_value=1000, value=10)

# Helper functions ----------------------------------------------------------------------
DATE_REGEX = re.compile(r'(\d{1,2}_\d{1,2}_\d{4})')  # finds dd_mm_yyyy with underscores

def extract_date_from_filename(filename: str):
    """
    Find last occurrence of dd_mm_yyyy in filename and return datetime.date or None.
    Accepts patterns like ..._01_01_2025.xlsx or ..._01_01_2025(1).csv
    """
    # take only the filename part
    name = filename.split("/")[-1].split("\\")[-1]
    matches = DATE_REGEX.findall(name)
    if not matches:
        return None
    date_str = matches[-1]  # last occurrence
    try:
        d = datetime.strptime(date_str, "%d_%m_%Y").date()
        return d
    except Exception:
        return None

def detect_awc_column(df: pd.DataFrame, hint: str = ""):
    """
    Try to detect AWC column in dataframe.
    Priority:
      1) if hint provided and column exists -> use it
      2) any column that case-insensitively contains 'awc'
      3) any column that is first non-numeric column
      4) fallback to first column
    Returns column name (string).
    """
    if hint:
        for col in df.columns:
            if col.strip().lower() == hint.strip().lower():
                return col
        # try partial match too
        for col in df.columns:
            if hint.strip().lower() in col.strip().lower():
                return col

    # look for 'awc' in column names
    for col in df.columns:
        if 'awc' in col.strip().lower():
            return col

    # prefer columns with object dtype and not fully numeric
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]) or not pd.api.types.is_numeric_dtype(df[col]):
            return col

    # fallback
    return df.columns[0]

# def read_table_from_upload(uploaded_file, skip_header=False):
#     """Read CSV or XLSX file from UploadedFile into pandas DataFrame."""
#     name = uploaded_file.name.lower()
#     uploaded_file.seek(0)

#     try:
#         # DAILY FILES → skip first 8 rows (start from row 9) if skip_header True
#         if skip_header:
#             if name.endswith(".csv"):
#                 return pd.read_csv(uploaded_file, skiprows=8)
#             else:
#                 return pd.read_excel(uploaded_file, sheet_name=0, skiprows=8)
#         else:
#             # AWC LIST FILE → do NOT skip rows
#             if name.endswith(".csv"):
#                 return pd.read_csv(uploaded_file)
#             else:
#                 return pd.read_excel(uploaded_file, sheet_name=0)
#     except Exception:
#         uploaded_file.seek(0)
#         if name.endswith(".csv"):
#             return pd.read_csv(uploaded_file, encoding="utf-8", engine="python")
#         raise



# st.title("AWC Multi-file Merger — combine multiple daily files by AWC")
# st.markdown("""
# Upload up to 100 CSV/XLSX files whose filenames contain a date in the format `dd_mm_yyyy`
# (or `dd_mm_yyyy(1)` etc). Also upload a single CSV/XLSX file that lists AWC names/codes (one column).
# The app will filter rows across all uploaded daily files for AWCs present in that list, add `Date`
# and `Day` columns (parsed from filename), and produce a combined file grouped by AWC.

# **After processing** the app will:
# - remove Sundays,
# - apply rules to `AWC didn't open`, `Total HCM Given`, `Morning Snack Given`,
# - color specific cells in the XLSX output (yellow/orange/pink/red),
# - produce a finalized XLSX or CSV for download.

# Note: This version will NOT overwrite existing "Morning Snack Given" values to 1.
# It keeps original snack values intact and only highlights them according to weekday & values.
# """)

# # --- Inputs -----------------------------------------------------------------------------
# col1, col2 = st.columns([2, 1])

# with col1:
#     st.subheader("1) Upload daily files (up to 100)")
#     uploaded_files = st.file_uploader(
#         "Choose CSV or XLSX files (multiple). Filenames must contain date like _dd_mm_yyyy",
#         type=["csv", "xlsx"],
#         accept_multiple_files=True,
#         help="Examples: file_name_ABC_01_01_2025.csv or file_name_DEF_01_01_2025(1).xlsx"
#     )

# with col2:
#     st.subheader("2) Upload AWC list file (single)")
#     awc_file = st.file_uploader(
#         "AWC list (csv or xlsx) — single file with one column containing AWC names/codes",
#         type=["csv", "xlsx"],
#         accept_multiple_files=False
#     )

# st.markdown("---")

# # Options
# st.sidebar.header("Options")
# awc_column_hint = st.sidebar.text_input("If you know the AWC column header, enter it (optional)", "")
# output_format = st.sidebar.selectbox("Output format", ["xlsx", "csv"])
# sort_by_date = st.sidebar.checkbox("Sort each AWC's rows by Date ascending", value=True)
# display_sample = st.sidebar.number_input("Show up to N matching rows in preview", min_value=0, max_value=1000, value=10)

# # Helper functions ----------------------------------------------------------------------
# DATE_REGEX = re.compile(r'(\d{1,2}_\d{1,2}_\d{4})')  # finds dd_mm_yyyy with underscores

# def extract_date_from_filename(filename: str):
#     """
#     Find last occurrence of dd_mm_yyyy in filename and return datetime.date or None.
#     Accepts patterns like ..._01_01_2025.xlsx or ..._01_01_2025(1).csv
#     """
#     # take only the filename part
#     name = filename.split("/")[-1].split("\\")[-1]
#     matches = DATE_REGEX.findall(name)
#     if not matches:
#         return None
#     date_str = matches[-1]  # last occurrence
#     try:
#         d = datetime.strptime(date_str, "%d_%m_%Y").date()
#         return d
#     except Exception:
#         return None

# def detect_awc_column(df: pd.DataFrame, hint: str = ""):
#     """
#     Try to detect AWC column in dataframe.
#     Priority:
#       1) if hint provided and column exists -> use it
#       2) any column that case-insensitively contains 'awc'
#       3) any column that is first non-numeric column
#       4) fallback to first column
#     Returns column name (string).
#     """
#     if hint:
#         for col in df.columns:
#             if col.strip().lower() == hint.strip().lower():
#                 return col
#         # try partial match too
#         for col in df.columns:
#             if hint.strip().lower() in col.strip().lower():
#                 return col

#     # look for 'awc' in column names
#     for col in df.columns:
#         if 'awc' in col.strip().lower():
#             return col

#     # prefer columns with object dtype and not fully numeric
#     for col in df.columns:
#         if pd.api.types.is_object_dtype(df[col]) or not pd.api.types.is_numeric_dtype(df[col]):
#             return col

#     # fallback
#     return df.columns[0]

def read_table_from_upload(uploaded_file, skip_header=False):
    """Read CSV or XLSX file from UploadedFile into pandas DataFrame."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)

    try:
        # DAILY FILES → skip first 8 rows (start from row 9) if skip_header True
        if skip_header:
            if name.endswith(".csv"):
                return pd.read_csv(uploaded_file, skiprows=8)
            else:
                return pd.read_excel(uploaded_file, sheet_name=0, skiprows=8)
        else:
            # AWC LIST FILE → do NOT skip rows
            if name.endswith(".csv"):
                return pd.read_csv(uploaded_file)
            else:
                return pd.read_excel(uploaded_file, sheet_name=0)
    except Exception:
        uploaded_file.seek(0)
        if name.endswith(".csv"):
            return pd.read_csv(uploaded_file, encoding="utf-8", engine="python")
        raise

def ensure_columns_exist(df: pd.DataFrame, cols_with_defaults: dict):
    """Ensure columns exist in df; if missing create with default value."""
    for c, default in cols_with_defaults.items():
        if c not in df.columns:
            df[c] = default
    return df

def apply_business_rules_and_highlights(df: pd.DataFrame):
    """
    Apply rules (remove Sundays, set flags/values) and return modified dataframe
    and a list of formatting instructions: list of tuples (row_idx, col_name, hex_fill).
    row_idx is integer index in df (0-based).
    
    Rules applied:
    0) Remove all Sunday rows
    1) If AWC didn't open = 1 (any day except Sunday) -> highlight "AWC didn't open" YELLOW
    2) If AWC didn't open = 0 AND Total HCM Given = 0 (any day except Sunday) -> highlight "Total HCM Given" ORANGE
    3) If AWC didn't open = 0 AND Morning Snack Given != 0 on Mon/Wed/Fri -> highlight "Morning Snack Given" PINK
    4) If AWC didn't open = 0 AND Morning Snack Given = 0 on Tue/Thu/Sat -> highlight "Morning Snack Given" RED
    """
    # work on a copy
    df2 = df.copy()

    # Ensure Date is datetime and Day is string weekday
    if 'Date' in df2.columns:
        df2['Date'] = pd.to_datetime(df2['Date'], errors='coerce')
        df2['Day'] = df2['Date'].dt.day_name()
    else:
        df2['Date'] = pd.NaT
        df2['Day'] = ''

    # RULE 0: Remove Sundays
    df2 = df2[df2['Day'] != 'Sunday'].reset_index(drop=True)

    # Ensure the key columns exist with numeric defaults
    df2 = ensure_columns_exist(df2, {
        "AWC didn't open": 0,
        "Total HCM Given": 0,
        "Morning Snack Given": 0
    })

    # convert these to numeric ints (coerce errors)
    for c in ["AWC didn't open", "Total HCM Given", "Morning Snack Given"]:
        df2[c] = pd.to_numeric(df2[c], errors='coerce').fillna(0).astype(int)

    # Prepare formatting list
    formats = []  # each entry: (row_index (0-based), column_name, hex_fill)

    # Colors (hex, openpyxl expects RRGGBB)
    YELLOW = "FFFF00"   # yellow
    ORANGE = "FFA500"   # orange
    PINK = "FFC0CB"     # pink
    RED = "FFC7CE"      # light red-ish

    # Iterate rows and apply highlighting rules
    for idx, row in df2.iterrows():
        day = str(row.get('Day', '')).strip()
        awc_didnt_open = int(row.get("AWC didn't open", 0))
        total_hcm = int(row.get('Total HCM Given', 0))
        snack = int(row.get('Morning Snack Given', 0))

        # RULE 1: AWC didn't open = 1 -> highlight "AWC didn't open" YELLOW
        if awc_didnt_open == 1:
            formats.append((idx, "AWC didn't open", YELLOW))
        
        # RULE 2: AWC didn't open = 0 AND Total HCM Given = 0 -> highlight "Total HCM Given" ORANGE
        if awc_didnt_open == 0 and total_hcm == 0:
            formats.append((idx, "Total HCM Given", ORANGE))
        
        # RULE 3: AWC didn't open = 0 AND Morning Snack Given != 0 on Mon/Wed/Fri -> highlight PINK
        if awc_didnt_open == 0 and day in ["Monday", "Wednesday", "Friday"] and snack != 0:
            formats.append((idx, "Morning Snack Given", PINK))
        
        # RULE 4: AWC didn't open = 0 AND Morning Snack Given = 0 on Tue/Thu/Sat -> highlight RED
        if awc_didnt_open == 0 and day in ["Tuesday", "Thursday", "Saturday"] and snack == 0:
            formats.append((idx, "Morning Snack Given", RED))

    # return modified df and formats
    return df2, formats

# Processing ---------------------------------------------------------------------------
if st.button("Process files"):
    if not uploaded_files or len(uploaded_files) == 0:
        st.error("Please upload at least one daily file.")
    elif not awc_file:
        st.error("Please upload the AWC list file.")
    else:
        st.info(f"Processing {len(uploaded_files)} files...")

        # Read AWC list
        try:
            awc_df = read_table_from_upload(awc_file, skip_header=False)
        except Exception as e:
            st.exception(f"Failed to read AWC file: {e}")
            st.stop()

        # Attempt to find the AWC column in list file
        try:
            awc_col_list = detect_awc_column(awc_df, awc_column_hint)
            awc_series = awc_df[awc_col_list].dropna().astype(str).str.strip()
            awc_set = set(awc_series.unique())
            if len(awc_set) == 0:
                st.warning("AWC file read successfully but no values detected in the chosen column. "
                           "Make sure the file has AWC entries.")
        except Exception as e:
            st.exception(f"Error detecting AWC column in the AWC file: {e}")
            st.stop()

        combined_rows = []
        found_awcs = set()
        file_process_errors = []

        for f in uploaded_files[:100]:  # limit to 100
            fname = f.name
            try:
                df = read_table_from_upload(f, skip_header=True)
            except Exception as e:
                file_process_errors.append((fname, str(e)))
                continue

            if df.shape[0] == 0:
                # nothing to do
                continue

            # Identify AWC column for this data file
            try:
                awc_col = detect_awc_column(df, awc_column_hint)
            except Exception:
                awc_col = df.columns[0]

            # Ensure column values are strings and stripped
            df[awc_col] = df[awc_col].astype(str).str.strip()

            # Filter rows where AWC in awc_set
            matched = df[df[awc_col].isin(awc_set)].copy()
            # record found AWCs
            found_awcs.update(matched[awc_col].unique())

            # Extract date from filename
            date_val = extract_date_from_filename(fname)
            if date_val is None:
                matched['__filename_date_missing'] = True
            else:
                matched['Date'] = pd.to_datetime(date_val)

            # Add Day column if Date available
            if 'Date' in matched.columns:
                matched['Day'] = matched['Date'].dt.day_name()
            else:
                matched['Day'] = ""

            # Add original filename column so user can trace origin
            matched['Source_Filename'] = fname

            if not matched.empty:
                combined_rows.append(matched)

        # Concatenate all matched rows
        if combined_rows:
            combined_df = pd.concat(combined_rows, ignore_index=True, sort=False)
        else:
            combined_df = pd.DataFrame()

        # Post-processing: check empty
        if combined_df.empty:
            st.warning("No rows matched the AWCs from the AWC list across uploaded files.")
            if file_process_errors:
                st.write("Some files could not be processed:")
                for fname, err in file_process_errors:
                    st.write(f"- {fname}: {err}")
            st.stop()

        # Normalize Date / Day
        if 'Date' in combined_df.columns:
            combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
        else:
            combined_df['Date'] = pd.NaT
        combined_df['Day'] = combined_df.get('Day', '').fillna('')
        # Normalize AWC column name in the combined df to 'AWC' for clarity
        detected_awc_col = None
        for col in combined_df.columns:
            try:
                vals = combined_df[col].dropna().astype(str)
                inter = set(vals.unique()).intersection(awc_set)
                if len(inter) > 0:
                    detected_awc_col = col
                    break
            except Exception:
                continue
        if detected_awc_col is None:
            detected_awc_col = combined_df.columns[0]
        if detected_awc_col != 'AWC':
            combined_df = combined_df.rename(columns={detected_awc_col: 'AWC'})

        # Optionally sort
        if sort_by_date:
            combined_df = combined_df.sort_values(by=['AWC', 'Date']).reset_index(drop=True)
        else:
            combined_df = combined_df.sort_values(by=['AWC']).reset_index(drop=True)

        # Reorder columns: AWC, Date, Day, Source_Filename, then rest
        rest_cols = [c for c in combined_df.columns if c not in ['AWC', 'Date', 'Day', 'Source_Filename']]
        final_cols = ['AWC', 'Date', 'Day', 'Source_Filename'] + rest_cols
        combined_df = combined_df[final_cols]

        # Apply the business rules & obtain formatting instructions
        finalized_df, formatting_instructions = apply_business_rules_and_highlights(combined_df)

        # Summary
        st.success("Processing finished and business rules applied.")
        st.write(f"Files processed: {len(uploaded_files)} (limited to first 100).")
        st.write(f"Total matched rows (after removing Sundays): {len(finalized_df)}")
        st.write(f"Number of unique AWCs in AWC list: {len(awc_set)}")
        st.write(f"Number of AWCs found across uploaded files: {len(found_awcs)}")
        not_found = set(awc_set) - set(found_awcs)
        st.write(f"AWCs from list not present in any file: {len(not_found)}")
        if len(not_found) <= 50:
            st.write(sorted(list(not_found)))
        else:
            st.write(f"(showing first 50) {sorted(list(not_found))[:50]}")

        if file_process_errors:
            st.warning("Some files failed to be read:")
            for fname, err in file_process_errors:
                st.write(f"- {fname}: {err}")

        # Preview
        st.subheader("Preview of finalized combined results (first rows)")
        if display_sample > 0:
            st.dataframe(finalized_df.head(display_sample))
        else:
            st.dataframe(finalized_df)

        # Prepare download
        buffer_name = f"combined_awc_data_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        if output_format == "xlsx":
            # STEP 1 — write dataframe normally using pandas (no formatting)
            temp_buffer = io.BytesIO()
            with pd.ExcelWriter(temp_buffer, engine="openpyxl") as writer:
                finalized_df.to_excel(writer, index=False, sheet_name="combined")

            # STEP 2 — reopen using openpyxl so that formatting persists
            temp_buffer.seek(0)
            from openpyxl import load_workbook
            wb = load_workbook(temp_buffer)
            ws = wb["combined"]

            # Build column index mapping
            col_to_idx = {col: idx+1 for idx, col in enumerate(finalized_df.columns)}

            # Create fills
            fills = {
                "FFFF00": PatternFill(start_color="FFFF00", fill_type="solid"),
                "FFA500": PatternFill(start_color="FFA500", fill_type="solid"),
                "FFC0CB": PatternFill(start_color="FFC0CB", fill_type="solid"),
                "FFC7CE": PatternFill(start_color="FFC7CE", fill_type="solid"),
            }

            # STEP 3 — Apply color formatting
            for (row_idx, col_name, hex_fill) in formatting_instructions:
                if col_name not in col_to_idx:
                    continue
                excel_row = row_idx + 2              # header offset
                excel_col = col_to_idx[col_name]
                cell = ws.cell(row=excel_row, column=excel_col)
                if hex_fill in fills:
                    cell.fill = fills[hex_fill]

            # STEP 4 — save final workbook into NEW buffer
            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.download_button(
                label="Download finalized results (XLSX with colors)",
                data=final_buffer,
                file_name=f"{buffer_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
                # except Exception as e:
                #     st.error(f"Failed to generate formatted XLSX: {e}")
                #     st.download_button(
                #         "Download CSV (fallback)",
                #         data=finalized_df.to_csv(index=False).encode("utf-8"),
                #         file_name=f"{buffer_name}.csv",
                #         mime="text/csv"
                #     )

        else:
            # CSV: colors cannot be preserved; deliver the updated numeric/flag values
            csv_bytes = finalized_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="Download finalized results (CSV)",
                data=csv_bytes,
                file_name=f"{buffer_name}.csv",
                mime="text/csv"
            )

        st.info("Notes:\n- Sundays were removed from the final file.\n- Coloring is applied to the XLSX download only.\n- This build will NOT overwrite Morning Snack Given to 1; original snack values are preserved and only highlighted according to weekday & values.")
